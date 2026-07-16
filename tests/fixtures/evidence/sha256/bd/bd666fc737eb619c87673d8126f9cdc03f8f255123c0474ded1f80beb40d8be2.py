"""Shared loaders + utilities for the principalandrepairersheets analysis.

Run from the principalandrepairersheets/ folder (the data dir) or via the
RAW env override. Every task script imports this.

Sources
-------
*.xls            legacy BIFF (Excel 2.x LABEL records) -> read with xlrd
*.xlsx           modern OOXML -> read with openpyxl
*.csv            contactseva_combined.csv (BOM, comma) -> csv module

TODAY is pinned to the project "current date" (2026-06-18) so the recency
bands are deterministic and reproducible, independent of wall-clock.
"""
from __future__ import annotations
import os, re, csv
from datetime import datetime, date
from collections import defaultdict

import openpyxl
import xlrd

# Data directory = parent of outputs/_scripts (i.e. the sheets folder).
HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.environ.get("RAW") or os.path.abspath(os.path.join(HERE, "..", ".."))

# Pinned "now" — the project current date. Bands are computed from this.
TODAY = date(2026, 6, 18)
CUTOFFS = {  # months -> first date that still counts as "used within"
    12: date(2025, 6, 18),
    24: date(2024, 6, 18),
    36: date(2023, 6, 18),
    48: date(2022, 6, 18),
}

# --------------------------------------------------------------------------
# UK postcode parsing
# --------------------------------------------------------------------------
_FULL = re.compile(r"^([A-Z]{1,2}[0-9][A-Z0-9]?)([0-9][A-Z]{2})$")
_OUT = re.compile(r"^([A-Z]{1,2}[0-9][A-Z0-9]?)$")


def parse_postcode(raw):
    """Return (full_canonical|None, outward|None, kind).

    kind in {'full','partial','none'}.
    'full'    -> a complete postcode; full_canonical='OL1 3QR', outward='OL1'.
    'partial' -> outward district only ('CH5'); full_canonical=None.
    'none'    -> not postcode-like.
    """
    if raw is None:
        return (None, None, "none")
    s = re.sub(r"[^A-Za-z0-9]", "", str(raw)).upper()
    if not s:
        return (None, None, "none")
    m = _FULL.fullmatch(s)
    if m:
        return (f"{m.group(1)} {m.group(2)}", m.group(1), "full")
    m = _OUT.fullmatch(s)
    if m:
        return (None, m.group(1), "partial")
    return (None, None, "none")


def find_postcode_in_fields(*fields):
    """Scan several free-text fields, return the best postcode found.

    Prefers a full postcode anywhere; else a partial. Returns
    (full|None, outward|None, kind). Used for contact/garage rows where the
    postcode may sit in any address column or be embedded in a blob.
    """
    blob = " ".join(str(f) for f in fields if f not in (None, ""))
    up = blob.upper()
    # full postcode embedded (with optional space)
    best_full = None
    for m in re.finditer(r"\b([A-Z]{1,2}[0-9][A-Z0-9]?)\s*([0-9][A-Z]{2})\b", up):
        best_full = (f"{m.group(1)} {m.group(2)}", m.group(1))
        break
    if best_full:
        return (best_full[0], best_full[1], "full")
    # bare outward token
    for m in re.finditer(r"\b([A-Z]{1,2}[0-9][A-Z0-9]?)\b", up):
        return (None, m.group(1), "partial")
    return (None, None, "none")


# --------------------------------------------------------------------------
# Name / code normalisation + fuzzy matching
# --------------------------------------------------------------------------
_NAME_STOP = {
    "ltd", "limited", "llp", "plc", "the", "co", "company", "uk", "and",
    "solicitors", "solicitor", "law", "legal", "lawyers", "advocates",
    "accident", "accidents", "repair", "repairs", "repairers", "repairer",
    "centre", "center", "services", "service", "garage", "garages",
    "autos", "auto", "motors", "motor", "body", "bodyshop", "bodyworks",
    "bodywork", "group", "claims", "claim", "management", "mgmt",
    "vehicle", "vehicles", "car", "cars", "crash", "coachworks", "ltd.",
    "associates", "partnership", "partners",
}


def norm_text(s):
    s = "" if s is None else str(s)
    s = s.lower().replace("&", " and ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def name_tokens(s, keep_stop=False):
    toks = norm_text(s).split()
    if keep_stop:
        return [t for t in toks if t]
    sig = [t for t in toks if t not in _NAME_STOP and len(t) > 1]
    return sig or [t for t in toks if t]  # fall back if everything was a stopword


def squash_name(s):
    """Significant tokens with stop-words removed, concatenated (no spaces).

    Collapses single-letter spacing so "C S Auto Repair Ltd" == "CS Auto Repair"
    (both -> 'cs', since auto/repair/ltd are stop-words).
    """
    toks = [t for t in norm_text(s).split() if t not in _NAME_STOP]
    if not toks:
        toks = norm_text(s).split()
    return "".join(toks)


def norm_code(s):
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def jaccard(a, b):
    a, b = set(a), set(b)
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def containment(a, b):
    a, b = set(a), set(b)
    if not a or not b:
        return 0.0
    return len(a & b) / min(len(a), len(b))


# --------------------------------------------------------------------------
# Loaders
# --------------------------------------------------------------------------
def _xls_rows(filename):
    """Yield dict rows from a BIFF .xls contact export.

    Header is on the 2nd row: Code,Name,Group,Address,City,County then an
    unlabelled postcode column. We capture the postcode as the best postcode
    found across the trailing address cells.
    """
    wb = xlrd.open_workbook(os.path.join(DATA, filename), logfile=open(os.devnull, "w"))
    sh = wb.sheet_by_index(0)
    # locate header row (the one whose first cell == 'Code')
    hdr_r = None
    for r in range(min(5, sh.nrows)):
        if str(sh.cell_value(r, 0)).strip().lower() == "code":
            hdr_r = r
            break
    if hdr_r is None:
        hdr_r = 1
    for r in range(hdr_r + 1, sh.nrows):
        vals = [sh.cell_value(r, c) for c in range(sh.ncols)]
        code = str(vals[0]).strip()
        if not code:
            continue
        name = str(vals[1]).strip() if len(vals) > 1 else ""
        group = str(vals[2]).strip() if len(vals) > 2 else ""
        tail = vals[3:]
        full, out, kind = find_postcode_in_fields(*tail)
        yield {
            "code": code, "name": name, "group": group,
            "address_blob": " ".join(str(v).strip() for v in tail if str(v).strip()),
            "full_pc": full, "outward": out, "pc_kind": kind,
            "_src": filename,
        }


def load_xls_contacts(filename):
    return list(_xls_rows(filename))


def load_repairer():
    return load_xls_contacts("REPAIRER.xls")


def load_contacts_csv():
    rows = []
    with open(os.path.join(DATA, "contactseva_combined.csv"), encoding="utf-8-sig",
              errors="replace") as f:
        for row in csv.DictReader(f):
            row = {(k or "").strip(): (v or "").strip() for k, v in row.items()}
            full, out, kind = find_postcode_in_fields(
                row.get("Postcode"), row.get("Address"), row.get("City"), row.get("County"))
            row["full_pc"], row["outward"], row["pc_kind"] = full, out, kind
            row["_src"] = "contactseva_combined.csv"
            rows.append(row)
    return rows


def load_providers_jobsheet():
    """Return (sheet1_rows, sheet2_rows)."""
    wb = openpyxl.load_workbook(os.path.join(DATA, "providersJOBSHEET.xlsx"), data_only=True)
    s1 = []
    ws = wb["Sheet1"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = row[0]
        if name is None or str(name).strip() == "":
            continue
        s1.append({
            "name": str(name).strip(),
            "code": str(row[1]).strip() if len(row) > 1 and row[1] else "",
            "inbox": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            "instructions": str(row[3]).strip() if len(row) > 3 and row[3] else "",
            "images_location": str(row[4]).strip() if len(row) > 4 and row[4] else "",
            "image_or_address": str(row[5]).strip() if len(row) > 5 and row[5] else "",
            "sending_report": str(row[6]).strip() if len(row) > 6 and row[6] else "",
        })
    s2 = []
    if "Sheet2" in wb.sheetnames:
        ws = wb["Sheet2"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            name = row[0]
            if name is None or str(name).strip() == "":
                continue
            s2.append({"name": str(name).strip(),
                       "cols": [str(c).strip() if c else "" for c in row]})
    wb.close()
    return s1, s2


def load_garages_jobsheet():
    wb = openpyxl.load_workbook(os.path.join(DATA, "garagesJOBSHEET.xlsx"), data_only=True)
    ws = wb["Sheet1"]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = row[0]
        if name is None or str(name).strip() == "":
            continue
        addr = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        email = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        phone = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        figures = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        full, o, kind = find_postcode_in_fields(addr)
        out.append({"name": str(name).strip(), "address": addr, "email": email,
                    "phone": phone, "figures": figures,
                    "full_pc": full, "outward": o, "pc_kind": kind})
    wb.close()
    return out


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def load_cases(filename):
    """Yield dict rows from everyrepairloc.xlsx or fulllist.xlsx.

    Keys: reference, engineer, registration, type, claimant, principal,
    date_created (date|None), inspected, reported, loc, inspection_type, claim_no.
    """
    wb = openpyxl.load_workbook(os.path.join(DATA, filename), read_only=True, data_only=True)
    ws = wb.active
    idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            idx = {str(h).strip(): j for j, h in enumerate(row) if h is not None}
            continue
        def g(col):
            j = idx.get(col)
            return row[j] if j is not None and j < len(row) else None
        yield {
            "reference": g("Reference"),
            "engineer": g("Engineer"),
            "registration": g("Registration"),
            "type": g("Type"),
            "claimant": g("Claimant"),
            "principal": (str(g("Principal")).strip() if g("Principal") is not None else ""),
            "date_created": _to_date(g("Date Created")),
            "inspected": _to_date(g("Inspected")),
            "reported": _to_date(g("Reported")),
            "loc": (str(g("Loc")).strip() if g("Loc") is not None else ""),
            "inspection_type": (str(g("Inspection Type")).strip() if g("Inspection Type") is not None else ""),
            "claim_no": g("Claim No"),
        }
    wb.close()


_PLACEHOLDER = re.compile(r"^\s*(f\.?a\.?o\.?[\s\-]*)?(the\s+)?court\s*$", re.I)


def is_placeholder_name(name):
    """True for EVA placeholder contact names like 'FAO The Court'."""
    n = norm_text(name)
    return (not n) or bool(_PLACEHOLDER.match(name or "")) or n in {
        "fao the court", "fao court", "the court", "court", "f a o the court"}


def derive_firm_from_address(addr):
    """Pull a firm name from an address blob when the Name is a placeholder.

    'C/O Zenith Lawyers 2 Strawberry Bank ...' -> 'Zenith Lawyers'
    'AA Motor Claims PO Box 585 ...'           -> 'AA Motor Claims'
    Stops at the first standalone number, 'PO Box', or a postcode token.
    """
    if not addr:
        return ""
    s = str(addr).strip()
    s = re.sub(r"^\s*c\s*/\s*o\s+", "", s, flags=re.I)        # drop leading C/O
    toks = s.replace(",", " ").split()
    out = []
    for t in toks:
        tl = t.lower().strip(".")
        if tl in ("po", "p.o") or re.match(r"\d", t) or _OUT.fullmatch(
                re.sub(r"[^A-Za-z0-9]", "", t).upper()):
            break
        out.append(t)
        if len(out) >= 5:
            break
    name = " ".join(out).strip(" ,-")
    return name if len(name) >= 3 else ""


def _good_name(name, addr):
    if is_placeholder_name(name):
        derived = derive_firm_from_address(addr)
        if derived:
            return derived, True
    return (name or ""), False


def build_code_name_map():
    """Map normalised EVA code -> {name, group, src, derived}.

    Priority: job-sheet provider names > CSV > xls exports. Placeholder names
    ('FAO The Court') are replaced with a firm name derived from the address,
    so real principals resolve to their real firm rather than the placeholder.
    """
    m = {}
    for fn in ["legal.xls", "REPAIRER.xls", "aALL.xls", "agent.xls", "broker.xls",
               "client.xls", "other.xls", "private.xls"]:
        try:
            for row in load_xls_contacts(fn):
                c = norm_code(row["code"])
                if not c:
                    continue
                name, derived = _good_name(row["name"], row["address_blob"])
                if name:
                    m[c] = {"name": name, "group": row["group"], "src": fn, "derived": derived}
        except Exception:
            pass
    for row in load_contacts_csv():
        c = norm_code(row.get("Code"))
        if not c:
            continue
        addr = " ".join(str(row.get(k, "")) for k in ("Address", "City", "County"))
        name, derived = _good_name(row.get("Name"), addr)
        if name:
            m[c] = {"name": name, "group": row.get("Group", ""), "src": "csv", "derived": derived}
    s1, _ = load_providers_jobsheet()
    for p in s1:
        c = norm_code(p["code"])
        if c and p["name"]:
            m[c] = {"name": p["name"], "group": "PROVIDER(jobsheet)", "src": "jobsheet",
                    "derived": False}
    return m


LOCKED_WRITES = []  # paths that were open elsewhere (e.g. Excel) and got a .new sibling


def _write_csv_to(path, header, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


def write_csv(path, header, rows):
    """Write a CSV. If `path` is locked (open in Excel), fall back to <path>.new
    instead of crashing, and record it in LOCKED_WRITES so the caller can warn.
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    import time
    for attempt in range(3):
        try:
            _write_csv_to(path, header, rows)
            return len(rows)
        except PermissionError:
            if attempt < 2:
                time.sleep(0.4)
    alt = path[:-4] + ".new.csv" if path.lower().endswith(".csv") else path + ".new"
    try:
        _write_csv_to(alt, header, rows)
    except PermissionError:
        pass
    LOCKED_WRITES.append(path)
    return len(rows)


def out_path(*parts):
    return os.path.join(DATA, "outputs", *parts)


if __name__ == "__main__":
    print("DATA =", DATA)
    print("TODAY =", TODAY, "CUTOFFS =", {k: str(v) for k, v in CUTOFFS.items()})
    reps = load_repairer()
    print("REPAIRER:", len(reps), "e.g.", reps[0])
    g = load_garages_jobsheet()
    print("garages:", len(g), "e.g.", g[0])
    s1, s2 = load_providers_jobsheet()
    print("providers sheet1:", len(s1), "sheet2:", len(s2))
    cmap = build_code_name_map()
    print("code->name entries:", len(cmap))
    # quick postcode self-test
    for t in ["OL13QR", "KY74AA", "CH5", "RH109NT", "M19", "EC1A1BB", "PO", "hello"]:
        print("  pc", t, "->", parse_postcode(t))
