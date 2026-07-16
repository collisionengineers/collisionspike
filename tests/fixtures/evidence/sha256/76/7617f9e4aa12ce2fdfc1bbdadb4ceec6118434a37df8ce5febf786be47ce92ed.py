import json
from pathlib import Path

import openpyxl

# raw/ is three levels up from this script (codexwork -> outputs -> principalandrepairersheets -> raw)
PATH = Path(__file__).resolve().parents[2] / "Backup of CE Job Sheet 260429.xlsm"
SHEET = "Principals"

# 1-based column map -> output key
COLS = {
    "name": 2,            # B Solicitor/Work Provider (name)
    "evaCode": 3,         # C EVA Code
    "boxCode": 4,         # D Box Code
    "inbox": 5,           # E Inbox
    "instructions": 6,    # F Solicitors Instructions
    "dragIntoEva": 7,     # G Drag in to EVA?
    # H (8) Sent Mino -> IGNORE
    "imagesLocation": 9,  # I Images location
    "modalityText": 10,   # J Image based or address
    "sendingReport": 11,  # K Sending Report
}


def norm(v):
    if v is None:
        return ""
    if isinstance(v, str):
        # preserve multi-line cell text, join physical line breaks with " \n "
        parts = v.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        parts = [p.strip() for p in parts]
        # drop empties produced by leading/trailing newlines but keep internal structure
        parts = [p for p in parts if p != ""]
        return " \n ".join(parts)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


wb = openpyxl.load_workbook(PATH, data_only=True, keep_vba=False)
ws = wb[SHEET]

# Row 1 is a blank/merged title row; the real column header is on row 2.
# Data therefore starts on row 3.
HEADER_ROW = 2
FIRST_DATA_ROW = 3
header = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, 13)]
print("HEADER ROW 2:", header)
print("max_row:", ws.max_row, "max_col:", ws.max_column)

providers = []
for r in range(FIRST_DATA_ROW, ws.max_row + 1):
    rec = {k: norm(ws.cell(row=r, column=c).value) for k, c in COLS.items()}
    # skip wholly blank rows (all mapped fields empty)
    if not any(rec.values()):
        continue
    providers.append(rec)

print("ROW COUNT (data):", len(providers))

out = {"providers": providers}
out_path = Path(__file__).resolve().parent / "principals_extracted.json"
with open(out_path, "w", encoding="utf-8") as fh:
    json.dump(out, fh, ensure_ascii=False, indent=2)
print("WROTE:", out_path)
print(json.dumps(out, ensure_ascii=False, indent=2))
